import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

from app_constants import AppConstant


class ExcelFileWorker:
    wb = None
    ws = None
    primary_column_idx = 0
    header_column_idx = 0
    file_name = None
    def __init__(self, file_name, primary_column_idx=0, header_column_idx=0):
        self.wb = openpyxl.open(file_name)
        self.ws = self.wb.active
        self.primary_column_idx = primary_column_idx
        self.header_column_idx = header_column_idx
        self.file_name = file_name

    def getNoOfRows(self):
        return self.ws.max_row

    def getNoOfColumns(self):
        return self.ws.max_column

    def getLastRow(self):
        return [col_.value for col_ in self.ws.iter_rows(self.ws.max_row, self.ws.max_row, 0, 0).__next__()]

    def getRowByIndex(self, row_idx):
        return [item.value for item in self.ws.iter_rows(row_idx + 1, row_idx + 1, 0, 0).__next__()]

    def getColumnByIndex(self, col_idx):
        return [item.value for item in self.ws.iter_cols(col_idx+1, col_idx+1, 0, 0).__next__()]

    def getHeaderLabels(self):
        return list(map(lambda itm: str(datetime.strftime(itm.date(), '%d-%m-%Y')) if isinstance(itm, datetime) else str(itm), self.getRowByIndex(self.header_column_idx)))

    def selectRowByFilter(self, include_only: list, row_idx):
        header_cols = self.getHeaderLabels()
        value_cols = self.ws.iter_rows(row_idx+1, row_idx+1, 0, 0).__next__()
        col_idx = 0
        for header_col, value_col in zip(header_cols, value_cols):
            if header_col in include_only:
                yield [value_col.value, (row_idx, col_idx)]  # returns value to the column as well as column's actual position in excel sheet.
            col_idx += 1

    def appendRow(self, data_):
        self.ws.append(data_)
        self.save()

    def deleteRowByIndex(self, row_idx):
        self.ws.delete_rows(row_idx+1)

    def updateByLocation(self, row_, column_, value_to_set, duplicacy_allowed=True):
        if not duplicacy_allowed:
            old_values = self.getColumnByIndex(column_)
            assert value_to_set not in old_values, f"'{value_to_set}' already exists."
            self.ws.cell(row_, column_, value_to_set)
        else:
            self.ws.cell(row_, column_, value_to_set)
        self.save()

    def save(self):
        self.wb.save(self.file_name)

def createStudentsExcelFile(file_name):
    # file_name: path of location where to save and file name to save.
    wb = Workbook()
    ws = wb.active

    headers_ = ["Admission No.", "Student Name", "Father Name", "Mobile No.", "Class", "Section", "Face Encoding File"]
    ws.append(headers_)
    for i in range(1, len(headers_)+1):
        ws.cell(1, i).font = Font(bold=True, size=12)
        ws.cell(1, i).fill = PatternFill(start_color="fcffde", end_color="fcffde", fill_type="solid")
        thin_style = Side(style='thin')
        ws.cell(1, i).border = Border(left=thin_style, right=thin_style, top=thin_style, bottom=thin_style)
    wb.save(file_name)

class AttendanceDBManager:
    CLASS_DIRECTORY = AppConstant.CLASS_DIRECTORY
    def __init__(self):
        pass

    def createThisYearAttendanceSheet(self, cls_sec_dir):
        path_to_save = os.path.join(self.CLASS_DIRECTORY, cls_sec_dir, f"{datetime.now().year}.xlsx")
        if os.path.exists(path_to_save):
            return
        wb = Workbook()
        ws = wb.active
        headers_ = ["Admission No.", "Student Name"]
        ws.append(headers_)
        for i in range(1, len(headers_) + 1):
            ws.cell(1, i).font = Font(bold=True, size=12)
            ws.cell(1, i).fill = PatternFill(start_color="fcffde", end_color="fcffde", fill_type="solid")
            thin_style = Side(style='thin')
            ws.cell(1, i).border = Border(left=thin_style, right=thin_style, top=thin_style, bottom=thin_style)
        wb.save(path_to_save)

    def insertNewStudent(self, cls_sec_dir, std_adm_no:int, std_name):
        attendance_sheet_path = os.path.join(self.CLASS_DIRECTORY, cls_sec_dir, f"{datetime.now().year}.xlsx")
        if os.path.exists(attendance_sheet_path):
            wb = openpyxl.open(attendance_sheet_path)
        else:
            wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([std_adm_no, std_name])
        wb.save(attendance_sheet_path)

def safe_max(seq):
    return max([int(item) for item in seq if str(item).isnumeric()] + [0])

if __name__ == '__main__':
    # createStudentsExcelFile(os.getcwd())
    # print(getNoOfRows("tests/File.xlsx"))
    # for item in getLastRow("tests/File.xlsx"):
    #     print(item)
    # print(getColumnByIndex("tests/File.xlsx", 1))
    # print(getColumnByIndex("tests/File.xlsx", 2))
    # print(safe_max([1, 2, "122", "012", "anc", "22a", "12"]))
    ed = ExcelFileWorker("tests/File.xlsx")
    ed.selectRowByFilter(["Name", "Class"], 1)
    # ed.updateByLocation(0, 0, "HELLO", True)
